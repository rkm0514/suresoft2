import openpyxl
import tkinter.filedialog
import tkinter as tk

# 기능사양분석서(FSA)
FSAfinder = tk.Tk()
FSAfinder.withdraw()
FSApath = tk.filedialog.askopenfilename(filetypes=[('Microsoft Excel File(.xlsx)', '.xlsx'), ('All files', '*')],
                                        title='Open FSA File',
                                        initialfile='*.xlsx')
FSAwb = openpyxl.load_workbook(FSApath)
FSAsht = FSAwb.active

# CAN DB
CANDBfinder = tk.Tk()
CANDBfinder.withdraw()
CANDBpath = tk.filedialog.askopenfilename(filetypes=[('Microsoft Excel File(.xlsx)', '.xlsx'), ('All files', '*')],
                                          title='Open CAN DB File',
                                          initialfile='*.xlsx')
CANDBwb = openpyxl.load_workbook(CANDBpath)
CANDBsht = CANDBwb.active


### CAN신호 검색, 저장 ##################################################################################################
for i in range(FSAsht.min_row + 2, FSAsht.max_row + 1):
    input_signal_list = FSAsht.cell(row=i, column=FSAsht.min_column + 11).value.splitlines()
    output_signal_list = FSAsht.cell(row=i, column=FSAsht.min_column + 12).value.splitlines()
    # Input CCP 읽고 CAN 불러오기
    for j in range(0, len(input_signal_list)):
        b_CAN_found = 0
        for k in range(CANDBsht.min_row + 2, CANDBsht.max_row + 1):
            # CCP와 매칭되는 CAN 검색
            if (CANDBsht.cell(row=k, column=CANDBsht.min_column + 12).value == input_signal_list[j]):
                b_CAN_found = 1
                if (j == 0):
                    FSAsht.cell(row=i, column=FSAsht.min_column + 8).value = 'CAN'
                    FSAsht.cell(row=i, column=FSAsht.min_column + 10).value = \
                        CANDBsht.cell(row=k, column=CANDBsht.min_column + 2).value
                else:
                    FSAsht.cell(row=i, column=FSAsht.min_column + 8).value += '\n' + 'CAN'
                    FSAsht.cell(row=i, column=FSAsht.min_column + 10).value += \
                        '\n' + CANDBsht.cell(row=k, column=CANDBsht.min_column + 2).value
                break
        # 매칭되는 CAN 신호 없을 시
        if (b_CAN_found == 0):
            if (j == 0):
                FSAsht.cell(row=i, column=FSAsht.min_column + 8).value = '확인 필요'
                FSAsht.cell(row=i, column=FSAsht.min_column + 10).value = '확인 필요'
            else:
                FSAsht.cell(row=i, column=FSAsht.min_column + 8).value += '\n' + '확인 필요'
                FSAsht.cell(row=i, column=FSAsht.min_column + 10).value += '\n' + '확인 필요'
    # Output CCP 읽고 CAN 불러오기
    for j in range(0, len(output_signal_list)):
        b_CAN_found = 0
        # CCP와 매칭되는 CAN 검색
        for k in range(CANDBsht.min_row + 2, CANDBsht.max_row + 1):
            if (CANDBsht.cell(row=k, column=CANDBsht.min_column + 12).value == output_signal_list[j]):
                b_CAN_found = 1
                if (j == 0):
                    FSAsht.cell(row=i, column=FSAsht.min_column + 13).value = 'CAN'
                    FSAsht.cell(row=i, column=FSAsht.min_column + 14).value = \
                        CANDBsht.cell(row=k, column=CANDBsht.min_column + 2).value
                else:
                    FSAsht.cell(row=i, column=FSAsht.min_column + 13).value += '\n' + 'CAN'
                    FSAsht.cell(row=i, column=FSAsht.min_column + 14).value += \
                        '\n' + CANDBsht.cell(row=k, column=CANDBsht.min_column + 2).value
                break
        # 매칭되는 CAN 신호 없을 시
        if (b_CAN_found == 0):
            if (j == 0):
                FSAsht.cell(row=i, column=FSAsht.min_column + 13).value = '확인 필요'
                FSAsht.cell(row=i, column=FSAsht.min_column + 14).value = '확인 필요'
            else:
                FSAsht.cell(row=i, column=FSAsht.min_column + 13).value += '\n' + '확인 필요'
                FSAsht.cell(row=i, column=FSAsht.min_column + 14).value += '\n' + '확인 필요'

########################################################################################################################

# 결과 파일 저장
FSAsaver = tk.Tk()
FSAsaver.withdraw()
FSApath = tkinter.filedialog.asksaveasfilename(filetypes=[('Microsoft Excel File(.xlsx)', '.xlsx'), ('All files', '*')],
                                               title='Save as..',
                                               initialfile='*.xlsx')
FSAwb.save(FSApath)